import os
import locale
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter


# CONFIGURAÇÕES 
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
diametro_de_borda = Side(border_style="medium", color="000000")
borda_grossa = Border(left=diametro_de_borda, right=diametro_de_borda, top=diametro_de_borda, bottom=diametro_de_borda)
thin = Side(style='thin')
medium = Side(style='medium')
detalhes_style = Font(size=8, italic=True, color="000000")



def aplicar_borda(ws, min_row, max_row, min_col, max_col):

    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(
                left=medium,
                right=medium ,
                top=medium if cell.row == min_row else thin,
                bottom=medium if cell.row == max_row else thin
            )

def criar_layout(ws):
    # TÍTULO
    ws.merge_cells('A1:J1')
    ws['A1'] = 'DESPESAS DEPTO PESSOAL - 2026 GRUPO PARANÁ'
    ws['A1'].font = Font(size=17, bold=True)
    

    # CABEÇALHO FIXO
    ws.merge_cells('A4:B4')
    ws['A4'].border = borda_grossa
    ws['A4'] = 'DESCRIÇÃO'

    
    #=======================
    # STYLE
    #========================
    
    #-- CINZA CLARA

    cinza_claro = PatternFill(start_color="D9D9D9", fill_type="solid")
    
    for i in [6,7,8,12,13,14,16]:
        ws[f'A{i}'].fill = cinza_claro
        ws[f'A{i}'].font = detalhes_style
        ws[f'A{i}'].alignment = Alignment(horizontal='right')


    ws['B16'].fill = cinza_claro
    ws['B17'].fill = cinza_claro
    ws['B18'].fill = cinza_claro
    ws['B19'].fill = cinza_claro


    #BORDAS

    aplicar_borda(ws, 5, 8, 1, 2)   # Salários
    aplicar_borda(ws, 9, 15, 1, 2)  # 13º
    aplicar_borda(ws, 20, 25, 1, 2) # Convênio
    aplicar_borda(ws, 16, 19, 1, 2) # Horas extras

    ws['A26'].border = borda_grossa

    ws['A16'].border = borda_grossa
    ws['A16'] = 'Horas\nExtras'
    ws['A16'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


    #-- TÍTULO
    header_fill = PatternFill(start_color="4F5D73", fill_type="solid")

    ws['A4'].font = Font(bold=True, color="FFFFFF")
    ws['A4'].fill = header_fill
    ws['A4'].alignment = Alignment(horizontal='center') 
    
    ws['A26'].font = Font(size=12, bold=True)

    # LARGURA
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10



    # ESTRUTURA FIXA
    estrutura = [
        ('Salarios',''),
        ('Bruto',''),
        ('Impostos',''),
        ('Desc. Func.',''),
        ('13º Sal.',''),
        ('Qtde Func.',''),
        ('Rescisões',''),
        ('Vale Transp.',''),
        ('Pago',''),
        ('Desc. Func.',''),
        ('Refeições desc. Fun.',''),
        ('Horas Extras','Desc.'),
        ('','60'),
        ('','100'),  
        ('','TOTAL'),  
        ('Convênio desc. Fun.',''),
        ('PM',''),
        ('Férias',''),
        ('Prêmio / Meta',''),
        ('Uniformes',''),
        ('Mat. Escritório',''),
        ('TOTAL DESPESAS','')
    ]

    for i,(nome, detalhamento) in enumerate(estrutura):

        if i in range(11, 15):
            ws[f'A{5 + i}'] = nome
            ws[f'B{5 + i}'] = detalhamento
        else:
            ws.merge_cells(f'A{5 + i}:B{5 + i}')
            ws[f'A{5 + i}'] = nome


    ws.merge_cells('A16:A19')

def pegar_proxima_coluna(mounth):

    colunas_por_mes = {
        'JAN': 3,
        'FEV': 5,
        'MAR': 7,
        'ABR': 9,
        'MAI': 11,
        'JUN': 13,
        'JUL': 15,
        'AGO': 17,
        'SET': 19,
        'OUT': 21,
        'NOV': 23,
        'DEZ': 25
    }
    return colunas_por_mes.get(mounth.upper(), 99)  


def escrever_mes(ws, mes, col):

    col_idx = column_index_from_string(col)
    col_prox = get_column_letter(col_idx + 1)

    ws.merge_cells(f'{col}4:{col_prox}4')
    ws[f'{col}4'] = mes

    for i in range(4,27):

        ws[f'{col}{i}'].alignment = Alignment(horizontal='center')
        ws[f'{col_prox}{i}'].alignment = Alignment(horizontal='center')

    header_fill = PatternFill(start_color="4F5D73", fill_type="solid")
    ws[f'{col}4'].font = Font(bold=True, color="FFFFFF")
    ws[f'{col}4'].fill = header_fill
    ws[f'{col}4'].alignment = Alignment(horizontal='center')
    ws[f'{col}4'].border = borda_grossa

    ws.column_dimensions[col].width = 14
    ws.column_dimensions[col_prox].width = 14

    for i in [6,7,8,12,13,14,16]:

        cinza_claro = PatternFill(start_color="D9D9D9", fill_type="solid")

        if i in (12,16,17,18,19):
            ws[f'{col}{i}'].fill = cinza_claro
            ws[f'{col_prox}{i}'].fill = cinza_claro


        ws[f'{col}{i}'].fill = cinza_claro
    

    ws[f'{col}16'].fill = cinza_claro
    ws[f'{col}17'].fill = cinza_claro
    ws[f'{col}18'].fill = cinza_claro
    ws[f'{col}19'].fill = cinza_claro


def preencher_dados(ws, dados,  col):

    col_idx = column_index_from_string(col)
    col_prox = get_column_letter(col_idx + 1)

    for i in range(5, 27):

        if i in (12,16,17,18,19):
            continue
        else: 
            ws.merge_cells(f'{col}{i}:{col_prox}{i}')

    ws[f'{col}6'] = locale.currency(dados['bruto_real'], grouping=True)
    ws[f'{col}8'] = locale.currency((dados['inss_planilha_custos'] + dados['inss_ferias']), grouping=True)
    ws[f'{col}10'] = dados['qtde_func']
    ws[f'{col}11'] = locale.currency(dados['rescisao_total'], grouping=True)
    ws[f'{col}12'] = 'Qtde'
    ws[f'{col_prox}12'] = dados['qtde_func_vt']
    ws[f'{col}13'] = locale.currency(dados['valor_vt'] , grouping=True)
    ws[f'{col}14'] = locale.currency(dados['vt_desc_func'], grouping=True)
    ws[f'{col}15'] = locale.currency(dados['refeicoes_desc_func'], grouping=True)
    ws[f'{col}16'] = 'Qtde'
    ws[f'{col_prox}16'] = 'Valor'

    # HORAS EXTRAS
    ws[f'{col}17'] = dados['quant_horas_extras_60']
    ws[f'{col_prox}17'] = locale.currency(dados['valor_horas_extras_60'], grouping=True)
    ws[f'{col}18'] = dados['quant_horas_extras_100']
    ws[f'{col_prox}18'] = locale.currency(dados['valor_horas_extras_100_com_dsr'], grouping=True)
    ws[f'{col}19'] = (dados['quant_horas_extras_60'] + dados['quant_horas_extras_100'])
    ws[f'{col_prox}19'] = locale.currency(
        dados['valor_horas_extras_60'] + dados['valor_horas_extras_100_com_dsr'],
        grouping=True
    )
    ws[f'{col}20'] = locale.currency((dados['valor_convenio_planilha_custos'] + dados['convenio_ferias']), grouping=True)

    ws[f'{col}22'] = locale.currency(dados['valor_ferias'], grouping=True)
    ws[f'{col}24'] = locale.currency(dados['valor_uniforme'], grouping=True)
    ws[f'{col}25'] = locale.currency(dados['valor_materiais'], grouping=True)

def aplicar_estilo_coluna_mes(ws, col):

    col_prox = col + 1

    aplicar_borda(ws, 5, 8, col, col_prox)   # Salários
    aplicar_borda(ws, 9, 15, col, col_prox)  # 13º
    aplicar_borda(ws, 20, 25, col, col_prox) # Convênio
    aplicar_borda(ws, 16, 19, col, col_prox) # Horas extras

def gerar_relatorio(dados, mes, ano, aba_nome="TOTAL GERAL"):

    ARQUIVO = f"RELATÓRIO DESPESAS REDE PARANÁ - {ano}.xlsx"

    # 1. ABRIR OU CRIAR O ARQUIVO
    if not os.path.exists(ARQUIVO):
        wb = Workbook()
        ws = wb.active
        ws.title = aba_nome
        criar_layout(ws)
    else:
        wb = load_workbook(ARQUIVO)
        
        # 2. VERIFICAR SE A ABA EXISTE
        if aba_nome in wb.sheetnames:
            ws = wb[aba_nome]
        else:
            ws = wb.create_sheet(title=aba_nome)
            criar_layout(ws)

    # 3. NOVA COLUNA
    col_index = pegar_proxima_coluna(mes)
    col_letra = get_column_letter(col_index)

    # 4. EVITA DUPLICAR MÊS (Otimizado)
    meses_existentes = [ws.cell(row=4, column=c).value for c in range(3, ws.max_column + 2, 2)]
    if mes in meses_existentes:
        print(f"Mês {mes} já existe na aba {aba_nome}!")
        return

    escrever_mes(ws, mes, col_letra)
    preencher_dados(ws, dados, col_letra)
    aplicar_estilo_coluna_mes(ws, col_index)

    wb.save(ARQUIVO)