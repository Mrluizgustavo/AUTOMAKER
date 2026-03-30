import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import numpy as np

from services.ui_theme import (
    ACCENT2,ACCENT, BG, BORDER, FONT_BODY, FONT_HEAD, FONT_MONO,
    FONT_SMALL, FONT_TITLE, SURFACE, SURFACE2,
    TEXT, TEXT_SUB, pill_button, _base_tela,
)

WARNING = "#F59E0B"


# ══════════════════════════════════════════════════════════════════════════════
# TELA DE ENTRADA
# ══════════════════════════════════════════════════════════════════════════════
def tela_dashboard_despesas(root):
    COR_DASH = "#0891B2"

    def corpo(root, content, cor):
        caminho = tk.StringVar(value="")

        # ── Card: arquivo ──
        card = tk.Frame(content, bg=SURFACE, padx=24, pady=20)
        card.pack(fill="x", pady=(0, 14))

        tk.Label(card, text="Arquivo de Relatório Excel", font=FONT_HEAD, bg=SURFACE, fg=TEXT).pack(anchor="w")
        tk.Label(card, text="Selecione o arquivo gerado pelo módulo Relatório Despesas (.xlsx)",
                 font=FONT_SMALL, bg=SURFACE, fg=TEXT_SUB).pack(anchor="w", pady=(2, 12))

        row = tk.Frame(card, bg=SURFACE)
        row.pack(fill="x")

        entry = tk.Entry(row, textvariable=caminho, bg=SURFACE2, fg=TEXT,
                         insertbackground=TEXT, relief="flat",
                         font=FONT_MONO, highlightthickness=1,
                         highlightbackground=BORDER, highlightcolor=cor)
        entry.pack(side="left", fill="x", expand=True, ipady=8, padx=(0, 10))
        entry.insert(0, "Nenhum arquivo selecionado...")

        def selecionar():
            p = filedialog.askopenfilename(
                title="Selecione o relatório de despesas",
                filetypes=[("Excel", "*.xlsx *.xls")]
            )
            if p:
                caminho.set(p)
                entry.config(fg=TEXT)

        pill_button(row, "Navegar", selecionar, color=cor, hover=ACCENT2).pack(side="right")

        # ── Card: aba ──
        card2 = tk.Frame(content, bg=SURFACE, padx=24, pady=16)
        card2.pack(fill="x", pady=(0, 14))
        tk.Label(card2, text="Aba / Unidade", font=FONT_HEAD, bg=SURFACE, fg=TEXT).pack(anchor="w", pady=(0, 8))

        aba_var = tk.StringVar(value="TOTAL GERAL")
        aba_row = tk.Frame(card2, bg=SURFACE)
        aba_row.pack(fill="x")
        tk.Label(aba_row, text="Aba:", bg=SURFACE, fg=TEXT_SUB, font=FONT_SMALL).pack(side="left", padx=(0, 8))
        aba_combo = ttk.Combobox(aba_row, textvariable=aba_var, state="readonly",
                                  font=FONT_BODY, width=30)
        aba_combo['values'] = ["TOTAL GERAL"]
        aba_combo.pack(side="left", fill="both", expand=True)

        def carregar_abas():
            arq = caminho.get()
            if not arq or "Nenhum" in arq:
                messagebox.showwarning("Atenção", "Selecione um arquivo primeiro.")
                return
            try:
                import openpyxl
                wb = openpyxl.load_workbook(arq, read_only=True, data_only=True)
                aba_combo['values'] = wb.sheetnames
                if wb.sheetnames:
                    aba_var.set(wb.sheetnames[0])
                wb.close()
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível ler as abas: {e}")

        pill_button(aba_row, "↻ Carregar Abas", carregar_abas,
                    color=SURFACE2, hover=BORDER, fg=TEXT_SUB).pack(side="left", padx=(10, 0))

        # ── Botões de ação ──
        btn_row = tk.Frame(content, bg=BG)
        btn_row.pack(fill="x", pady=(4, 0))

        def abrir_dashboard():
            arq = caminho.get()
            if not arq or "Nenhum" in arq:
                messagebox.showwarning("Atenção", "Selecione um arquivo antes de continuar.")
                return
            aba = aba_var.get()
            try:
                dados = _ler_dados_relatorio(arq, aba)
                if not dados:
                    messagebox.showwarning("Sem dados", "Não foram encontrados dados de meses nesta aba.")
                    return
                _janela_dashboard(root, dados, aba, arq, cor)
            except Exception as e:
                messagebox.showerror("Erro ao carregar", f"Falha ao processar o arquivo:\n{e}")

        def abrir_comparativo():
            arq = caminho.get()
            if not arq or "Nenhum" in arq:
                messagebox.showwarning("Atenção", "Selecione um arquivo antes de continuar.")
                return
            try:
                import openpyxl
                wb = openpyxl.load_workbook(arq, read_only=True, data_only=True)
                abas = wb.sheetnames
                wb.close()
                if len(abas) < 2:
                    messagebox.showinfo("Info", "São necessárias pelo menos 2 abas para o comparativo.")
                    return
                _janela_comparativo_lojas(root, arq, abas, cor)
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível abrir o arquivo:\n{e}")

        pill_button(btn_row, "📈  Abrir Dashboard", abrir_dashboard,
                    color=cor, hover=ACCENT2).pack(side="left", fill="x", expand=True, ipady=4, padx=(0, 6))
        pill_button(btn_row, "🏪  Comparativo por Loja", abrir_comparativo,
                    color="#065F46", hover="#047857").pack(side="left", fill="x", expand=True, ipady=4)

    _base_tela(root, "Dashboard Despesas", "📈", COR_DASH, corpo)


# ══════════════════════════════════════════════════════════════════════════════
# LEITURA DO EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def _ler_dados_relatorio(caminho, aba):
    import openpyxl

    LINHA_MAP = {
        6:  "bruto",
        8:  "impostos",
        10: "qtde_func",
        11: "rescisao",
        13: "valor_vt",
        14: "vt_desc_func",
        15: "refeicoes_desc",
        17: "he60_qtde",
        18: "he100_qtde",
        19: "he_total_qtde",
        20: "convenio",
        21: "convenio_ferias",
        22: "ferias",
        24: "uniformes",
        25: "materiais",
    }

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba]
    dados_meses = []
    max_col = ws.max_column or 30

    def parse_currency(val):
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).replace("R$", "").replace(".", "").replace(",", ".").replace(" ", "").strip()
        try:
            return float(s)
        except Exception:
            return 0.0

    def parse_int(val):
        if val is None:
            return 0
        if isinstance(val, (int, float)):
            return int(val)
        try:
            return int(str(val).strip())
        except Exception:
            return 0

    for col in range(3, max_col + 1, 2):
        mes_header = ws.cell(row=4, column=col).value
        if not mes_header or not isinstance(mes_header, str):
            continue
        mes_header = str(mes_header).strip().upper()
        if not mes_header:
            continue

        d = {"mes": mes_header}
        for linha, chave in LINHA_MAP.items():
            raw = ws.cell(row=linha, column=col).value
            if chave in ("qtde_func", "qtde_func_vt", "he60_qtde", "he100_qtde", "he_total_qtde"):
                d[chave] = parse_int(raw)
            else:
                d[chave] = parse_currency(raw)

        d["he60_valor"]     = parse_currency(ws.cell(row=17, column=col + 1).value)
        d["he100_valor"]    = parse_currency(ws.cell(row=18, column=col + 1).value)
        d["he_total_valor"] = parse_currency(ws.cell(row=19, column=col + 1).value)
        d["qtde_func_vt"]   = parse_int(ws.cell(row=12, column=col + 1).value)

        dados_meses.append(d)

    wb.close()
    return dados_meses


def _ler_ultimo_mes_aba(caminho, aba):
    dados = _ler_dados_relatorio(caminho, aba)
    return dados[-1] if dados else None


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS MATPLOTLIB COMPARTILHADOS
# ══════════════════════════════════════════════════════════════════════════════
BG_M   = "#12121A"
FG_M   = "#F1F0FF"
SUB_M  = "#8884A8"
COLORS = ["#7C3AED", "#0891B2", "#10B981", "#F59E0B", "#EF4444",
          "#A855F7", "#06B6D4", "#34D399", "#FBBF24", "#F87171"]

MPL_STYLE = {
    "figure.facecolor": BG_M, "axes.facecolor": "#1C1C28",
    "axes.edgecolor": "#2A2A3E", "axes.labelcolor": FG_M,
    "text.color": FG_M, "xtick.color": SUB_M, "ytick.color": SUB_M,
    "grid.color": "#2A2A3E", "grid.alpha": 0.6,
    "font.family": "DejaVu Sans",
}

def fmt_brl(v):
    return f"R$ {v:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")

def _setup_mpl():
    import matplotlib
    matplotlib.use("TkAgg")
    import matplotlib.pyplot as plt
    plt.rcParams.update(MPL_STYLE)
    return plt

def _embed(fig, parent, plt):
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    c = FigureCanvasTkAgg(fig, master=parent)
    c.draw()
    c.get_tk_widget().pack(fill="both", expand=True)
    plt.close(fig)

def _card_frame(parent, titulo, subtitulo="", accent_color="#0891B2"):
    f = tk.Frame(parent, bg=SURFACE)
    h = tk.Frame(f, bg=SURFACE, padx=14, pady=10)
    h.pack(fill="x")
    tk.Label(h, text=titulo, font=FONT_HEAD, bg=SURFACE, fg=TEXT).pack(anchor="w")
    if subtitulo:
        tk.Label(h, text=subtitulo, font=FONT_SMALL, bg=SURFACE, fg=TEXT_SUB).pack(anchor="w")
    tk.Frame(f, bg=accent_color, height=2).pack(fill="x")
    return f

def _make_fig(plt, w_in, h_px):
    return plt.figure(figsize=(w_in, h_px / 96), dpi=96)

def _scroll_window(win):
    
    outer = tk.Frame(win, bg=BG)
    outer.pack(fill="both", expand=True)
    
    cs = tk.Canvas(outer, bg=BG, highlightthickness=0)
    sb = ttk.Scrollbar(outer, orient="vertical", command=cs.yview)
    cs.configure(yscrollcommand=sb.set)
    
    sb.pack(side="right", fill="y")
    cs.pack(side="left", fill="both", expand=True)
    
    inner = tk.Frame(cs, bg=BG)
    
    frame_id = cs.create_window((0, 0), window=inner, anchor="nw")
    
    inner.bind("<Configure>", lambda e: cs.configure(scrollregion=cs.bbox("all")))
    
    cs.bind("<Configure>", lambda e: cs.itemconfig(frame_id, width=e.width))
    
    cs.bind_all("<MouseWheel>", lambda e: cs.yview_scroll(int(-1 * (e.delta / 120)), "units"))
    
    return inner


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD PRINCIPAL — aba única com layout mix
# ══════════════════════════════════════════════════════════════════════════════
def _janela_dashboard(root, dados, aba_nome, caminho_arq, cor):
    try:
        plt = _setup_mpl()
    except ImportError:
        messagebox.showerror("Dependência faltando",
                             "matplotlib e numpy são necessários.\n"
                             "Instale com: pip install matplotlib numpy")
        return

    win = tk.Toplevel(root)
    win.title(f"Dashboard — {aba_nome}")
    win.configure(bg=BG)
    win.geometry("1440x920")

    tk.Frame(win, bg=cor, height=5).pack(fill="x")
    top = tk.Frame(win, bg=BG)
    top.pack(fill="x", padx=28, pady=(14, 8))
    tk.Label(top, text="📈", font=("Segoe UI", 22), bg=BG, fg=cor).pack(side="left", padx=(0, 10))
    col_t = tk.Frame(top, bg=BG)
    col_t.pack(side="left")
    tk.Label(col_t, text=f"Dashboard Despesas — {aba_nome}", font=FONT_TITLE, bg=BG, fg=TEXT).pack(anchor="w")
    tk.Label(col_t, text=f"{len(dados)} período(s) carregados",
             font=FONT_SMALL, bg=BG, fg=TEXT_SUB).pack(anchor="w")

    meses = [d["mes"] for d in dados]
    inner = _scroll_window(win)

    # ──────────────────────────────────────────────────
    # LINHA 0 — KPI chips (5 métricas do último mês)
    # ──────────────────────────────────────────────────

    ultimo = dados[-1]

    total_folha = sum(v for k, v in ultimo.items() if isinstance(v, float) and "qtde" not in k)

    kpis_chips = [
        ("💰 Folha Total",  fmt_brl(total_folha),                                COLORS[0]),
        ("👥 Funcionários", str(ultimo.get("qtde_func", 0)),                     COLORS[1]),
        ("⏰ Custo HE",     fmt_brl(ultimo.get("he_total_valor", 0)),            COLORS[4]),
        ("🏥 Convênio",     fmt_brl(ultimo.get("convenio", 0)),                  COLORS[2]),
        ("🚌 VT Líquido",   fmt_brl(max(ultimo.get("valor_vt", 0)
                                        - ultimo.get("vt_desc_func", 0), 0)),    COLORS[3]),
    ]
    chips_row = tk.Frame(inner, bg=BG)
    chips_row.pack(fill="x", padx=16, pady=(10, 4))
    for label, valor, clr in kpis_chips:
        chip = tk.Frame(chips_row, bg=SURFACE, padx=16, pady=12)
        chip.pack(side="left", fill="x", expand=True, padx=6)
        tk.Frame(chip, bg=clr, height=3).pack(fill="x")
        tk.Label(chip, text=valor, font=("Segoe UI", 14, "bold"),
                 bg=SURFACE, fg=TEXT).pack(pady=(8, 2))
        tk.Label(chip, text=label, font=FONT_SMALL, bg=SURFACE, fg=TEXT_SUB).pack()

    # ──────────────────────────────────────────────────
    # LINHA 1 — Visão Executiva (larga) + Rosca último mês
    # ──────────────────────────────────────────────────
    row1 = tk.Frame(inner, bg=BG)
    row1.pack(fill="x", padx=16, pady=6)

    c1a = _card_frame(row1, "1. Visão Executiva da Folha",
                      "Salários + Benefícios + Variáveis + Adm por período", COLORS[0])
    c1a.pack(side="left", fill="both", expand=True, padx=(0, 6))

    def kpi_visao_exec():
        salarios   = [d.get("bruto", 0) + d.get("impostos", 0) for d in dados]
        beneficios = [d.get("convenio", 0)
                      + max(d.get("valor_vt", 0) - d.get("vt_desc_func", 0), 0)
                      for d in dados]
        variaveis  = [d.get("rescisao", 0) + d.get("he_total_valor", 0) + d.get("ferias", 0)
                      for d in dados]
        outros     = [d.get("uniformes", 0) + d.get("materiais", 0) for d in dados]
        x = np.arange(len(meses))
        fig = _make_fig(plt, 8.5, 300)
        ax = fig.add_subplot(111)
        ax.bar(x, salarios,   label="Salários + Impostos", color=COLORS[0], alpha=0.92)
        ax.bar(x, beneficios, label="Benefícios Líquidos", color=COLORS[1], alpha=0.92,
               bottom=salarios)
        b2 = [a + b for a, b in zip(salarios, beneficios)]
        ax.bar(x, variaveis, label="Variáveis", color=COLORS[4], alpha=0.92, bottom=b2)
        b3 = [a + b for a, b in zip(b2, variaveis)]
        ax.bar(x, outros, label="Adm/Outros", color=COLORS[3], alpha=0.92, bottom=b3)
        ax.set_xticks(x)
        ax.set_xticklabels(meses, fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: fmt_brl(v)))
        ax.tick_params(axis='y', labelsize=7)
        ax.legend(fontsize=7.5, facecolor=BG_M, edgecolor=BG_M,
                  labelcolor=FG_M, loc="upper left")
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c1a, plt)

    kpi_visao_exec()

    c1b = _card_frame(row1, "2. Composição — Último Mês",
                      f"Rosca com proporção de cada bloco — {meses[-1]}", COLORS[2])
    c1b.pack(side="left", fill="both", expand=True)


    def kpi_rosca():
        ul = dados[-1]
        labels = ["Sal. Bruto", "Rescisões", "Férias", "HE", "Convênio", "VT", "Adm"]
        valores = [
            ul.get("bruto", 0), ul.get("rescisao", 0), ul.get("ferias", 0),
            ul.get("he_total_valor", 0), ul.get("convenio", 0),
            max(ul.get("valor_vt", 0) - ul.get("vt_desc_func", 0), 0),
            ul.get("uniformes", 0) + ul.get("materiais", 0),
        ]
        valores = [max(v, 0) for v in valores]
        total_v = sum(valores) or 1
        fig = _make_fig(plt, 3.8, 300)
        ax = fig.add_subplot(111)
        wedges, _ = ax.pie(valores, startangle=90, colors=COLORS[:7],
                           wedgeprops=dict(width=0.6), explode=[0.04] * 7)
        ax.text(0, 0, fmt_brl(total_v), ha="center", va="center",
                fontsize=9, color=FG_M, fontweight="bold")
        ax.legend(wedges,
                  [f"{l}  {(v / total_v * 100):.0f}%" for l, v in zip(labels, valores)],
                  fontsize=7, loc="lower center", bbox_to_anchor=(0.5, -0.18),
                  facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M, ncol=2)
        fig.tight_layout(pad=1.2)
        _embed(fig, c1b, plt)

    kpi_rosca()

    # ──────────────────────────────────────────────────
    # LINHA 2 — Área suavizada evolução + HE barras
    # ──────────────────────────────────────────────────
    row2 = tk.Frame(inner, bg=BG)
    row2.pack(fill="x", padx=16, pady=6)

    c2a = _card_frame(row2, "3. Evolução da Folha — Área Suavizada",
                      "Bruto vs Total Despesas com interpolação spline", COLORS[1])
    c2a.pack(side="left", fill="both", expand=True, padx=(0, 6))

    def kpi_area_folha():
        x_idx = np.arange(len(meses))
        bruto = [d.get("bruto", 0) for d in dados]
        total = [sum(v for k, v in d.items() if isinstance(v, float) and "qtde" not in k)
                 for d in dados]
        fig = _make_fig(plt, 6.5, 280)
        ax = fig.add_subplot(111)
        try:
            from scipy.interpolate import make_interp_spline    
            if len(meses) >= 4:
                xn    = np.linspace(0, len(meses) - 1, 300)
                spl_b = make_interp_spline(x_idx, bruto, k=3)(xn)
                spl_t = make_interp_spline(x_idx, total, k=3)(xn)
                ax.fill_between(xn, spl_t, alpha=0.15, color=COLORS[4])
                ax.fill_between(xn, spl_b, alpha=0.22, color=COLORS[1])
                ax.plot(xn, spl_t, color=COLORS[4], linewidth=2,   label="Total Despesas")
                ax.plot(xn, spl_b, color=COLORS[1], linewidth=2,   label="Salário Bruto")
            else:
                raise ImportError
        except ImportError:
            ax.fill_between(x_idx, total, alpha=0.15, color=COLORS[4])
            ax.fill_between(x_idx, bruto, alpha=0.22, color=COLORS[1])
            ax.plot(x_idx, total, color=COLORS[4], linewidth=2, marker="o", label="Total Despesas")
            ax.plot(x_idx, bruto, color=COLORS[1], linewidth=2, marker="o", label="Salário Bruto")
        ax.scatter(x_idx, bruto, color=COLORS[1], zorder=5, s=40)
        ax.scatter(x_idx, total, color=COLORS[4], zorder=5, s=40)
        ax.set_xticks(x_idx)
        ax.set_xticklabels(meses, fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: fmt_brl(v)))
        ax.tick_params(axis='y', labelsize=7)
        ax.legend(fontsize=8, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M)
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c2a, plt)

    kpi_area_folha()

    c2b = _card_frame(row2, "4. Termômetro de Horas Extras",
                      "HE 60% vs HE 100% por período", COLORS[4])
    c2b.pack(side="left", fill="both", expand=True )
    

    def kpi_he():
        he60  = [d.get("he60_valor", 0)  for d in dados]
        he100 = [d.get("he100_valor", 0) for d in dados]
        x = np.arange(len(meses))
        w = 0.35
        fig = _make_fig(plt, 4.2, 280)
        ax = fig.add_subplot(111)
        ax.bar(x - w / 2, he60,  w, label="HE 60%",  color=COLORS[1], alpha=0.9)
        ax.bar(x + w / 2, he100, w, label="HE 100%", color=COLORS[4], alpha=0.9)
        ax.set_xticks(x)
        ax.set_xticklabels(meses, fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: fmt_brl(v)))
        ax.tick_params(axis='y', labelsize=7)
        ax.legend(fontsize=8, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M)
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c2b, plt)

    kpi_he()

    # ──────────────────────────────────────────────────
    # LINHA 3 — Radar saúde DP + Custo por cabeça área
    # ──────────────────────────────────────────────────
    row3 = tk.Frame(inner, bg=BG)
    row3.pack(fill="x", padx=16, pady=6)

    c3a = _card_frame(row3, "5. Radar de Saúde do DP",
                      "Visão multidimensional — último mês vs meta de referência", COLORS[5])
    c3a.pack(side="left", fill="both", expand=True)


    def kpi_radar():
        ul  = dados[-1]
        b   = max(ul.get("bruto", 1), 1)
        q   = max(ul.get("qtde_func", 1), 1)
        tot = sum(v for k, v in ul.items() if isinstance(v, float) and "qtde" not in k) or 1
        dims = ["Turnover\nCtrl", "Efic. HE", "Recup.\nDesc.", "Custo\nBenef.", "Adm\nCtrl"]
        vals = [
            max(0, 100 - (ul.get("rescisao", 0) / b * 100) * 5),
            max(0, 100 - (ul.get("he_total_qtde", 0) / q) * 80),
            min(100, (ul.get("vt_desc_func", 0) + ul.get("refeicoes_desc", 0)) / b * 1000),
            min(100, max(0, 100 - (ul.get("convenio", 0) + ul.get("valor_vt", 0)) / tot * 200)),
            min(100, max(0, 100 - (ul.get("uniformes", 0) + ul.get("materiais", 0)) / tot * 300)),
        ]
        metas = [70, 75, 60, 65, 80]
        N      = len(dims)
        angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist() + [0]
        vals_p  = vals + vals[:1]
        metas_p = metas + metas[:1]
        fig = _make_fig(plt, 4.6, 300)
        ax  = fig.add_subplot(111, polar=True)
        ax.set_facecolor("#1C1C28")
        ax.plot(angles, metas_p, color=WARNING, linewidth=1.2,
                linestyle="--", alpha=0.7, label="Meta")
        ax.fill(angles, vals_p, color=COLORS[5], alpha=0.25)
        ax.plot(angles, vals_p, color=COLORS[5], linewidth=2,
                marker="o", markersize=5, label="Atual")
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(dims, fontsize=8, color=FG_M)
        ax.set_ylim(0, 100)
        ax.set_yticks([25, 50, 75, 100])
        ax.set_yticklabels(["25", "50", "75", "100"], fontsize=6, color=SUB_M)
        ax.spines["polar"].set_color("#2A2A3E")
        ax.grid(color="#2A2A3E", alpha=0.7)
        ax.legend(fontsize=8, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M,
                  loc="upper right", bbox_to_anchor=(1.35, 1.1))
        fig.patch.set_facecolor(BG_M)
        fig.tight_layout(pad=1.5)
        _embed(fig, c3a, plt)

    kpi_radar()

    c3b = _card_frame(row3, "6. Custo por Cabeça — Tendência",
                      "Total ÷ funcionários com rótulos e área suavizada", COLORS[2])
    c3b.pack(side="left", fill="both", expand=True)

    def kpi_cpc_area():
        x_idx = np.arange(len(meses))
        total = [sum(v for k, v in d.items() if isinstance(v, float) and "qtde" not in k)
                 for d in dados]
        qtde  = [max(d.get("qtde_func", 1), 1) for d in dados]
        cpc   = [t / q for t, q in zip(total, qtde)]
        fig = _make_fig(plt, 6.2, 300)
        ax  = fig.add_subplot(111)
        try:
            from scipy.interpolate import make_interp_spline
            if len(meses) >= 4:
                xn  = np.linspace(0, len(meses) - 1, 300)
                spl = make_interp_spline(x_idx, cpc, k=3)(xn)
                ax.fill_between(xn, spl, alpha=0.2, color=COLORS[2])
                ax.plot(xn, spl, color=COLORS[2], linewidth=2.5)
            else:
                raise ImportError
        except ImportError:
            ax.fill_between(x_idx, cpc, alpha=0.2, color=COLORS[2])
            ax.plot(x_idx, cpc, color=COLORS[2], linewidth=2.5, marker="o")
        ax.scatter(x_idx, cpc, color=COLORS[2], zorder=5, s=50,
                   edgecolors=BG_M, linewidths=1.5)
        for xi, yi in zip(x_idx, cpc):
            ax.text(xi, yi + max(cpc) * 0.03, fmt_brl(yi),
                    ha="center", fontsize=7, color=FG_M)
        ax.set_xticks(x_idx)
        ax.set_xticklabels(meses, fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: fmt_brl(v)))
        ax.tick_params(axis='y', labelsize=7)
        ax.set_ylabel("R$ / funcionário", fontsize=8)
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c3b, plt)

    kpi_cpc_area()

    # ──────────────────────────────────────────────────
    # LINHA 4 — 3 colunas: Turnover | Benefícios | Recuperação
    # ──────────────────────────────────────────────────
    row4 = tk.Frame(inner, bg=BG)
    row4.pack(fill="x", padx=16, pady=6)

    c4a = _card_frame(row4, "7. Sangria de Turnover",
                      "Rescisões / Bruto — crítico acima de 10%", COLORS[4])
    c4a.pack(side="left", fill="both", expand=True, padx=(0, 6))

    def kpi_turnover():
        ratio = [(d.get("rescisao", 0) / max(d.get("bruto", 1), 1)) * 100 for d in dados]
        cores = [COLORS[4] if r > 10 else COLORS[2] for r in ratio]
        fig = _make_fig(plt, 4.0, 260)
        ax  = fig.add_subplot(111)
        bars = ax.bar(meses, ratio, color=cores, alpha=0.9)
        ax.axhline(10, color=WARNING, linestyle="--", linewidth=1.5, label="Limite 10%")
        for bar, val in zip(bars, ratio):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 0.2, f"{val:.1f}%",
                    ha="center", fontsize=7, color=FG_M)
        ax.set_xticklabels(meses, fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.1f}%"))
        ax.tick_params(axis='y', labelsize=7)
        ax.legend(fontsize=8, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M)
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c4a, plt)

    kpi_turnover()

    c4b = _card_frame(row4, "8. VT + Convênio",
                      "Bruto vs desconto do funcionário + subsídio real", COLORS[1])
    c4b.pack(side="left", fill="both", expand=True, padx=(0, 6))

    def kpi_beneficios():
        vt_b = [d.get("valor_vt", 0)    for d in dados]
        vt_d = [d.get("vt_desc_func", 0) for d in dados]
        conv = [d.get("convenio", 0)     for d in dados]
        sub  = [max(b - desc + c, 0) for b, desc, c in zip(vt_b, vt_d, conv)]
        fig = _make_fig(plt, 4.0, 260)
        ax  = fig.add_subplot(111)
        ax.bar(meses, vt_b,  label="VT Bruto",    color=COLORS[0], alpha=0.85)
        ax.bar(meses, vt_d,  label="Desc. Func.", color=COLORS[3], alpha=0.85)
        ax.plot(meses, sub, marker="s", color=COLORS[2],
                linewidth=2, label="Subsídio Real", markersize=6)
        ax.set_xticklabels(meses, fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: fmt_brl(v)))
        ax.tick_params(axis='y', labelsize=7)
        ax.legend(fontsize=7.5, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M)
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c4b, plt)

    kpi_beneficios()

    c4c = _card_frame(row4, "9. Taxa de Recuperação",
                      "Descontos do funcionário / Bruto", COLORS[5])
    c4c.pack(side="left", fill="both", expand=True)

    def kpi_recuperacao():
        rec = [(d.get("vt_desc_func", 0) + d.get("refeicoes_desc", 0))
               / max(d.get("bruto", 1), 1) * 100 for d in dados]
        x_idx = np.arange(len(meses))
        fig = _make_fig(plt, 4.0, 260)
        ax  = fig.add_subplot(111)
        ax.fill_between(x_idx, rec, alpha=0.2, color=COLORS[5])
        ax.plot(x_idx, rec, marker="D", color=COLORS[5], linewidth=2.5, markersize=6)
        ax.set_xticks(x_idx)
        ax.set_xticklabels(meses, fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.2f}%"))
        ax.tick_params(axis='y', labelsize=7)
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c4c, plt)

    kpi_recuperacao()

    # ──────────────────────────────────────────────────
    # LINHA 5 — Adm + HE/Func (2 colunas)
    # ──────────────────────────────────────────────────
    row5 = tk.Frame(inner, bg=BG)
    row5.pack(fill="x", padx=16, pady=6)

    c5a = _card_frame(row5, "10. Custos Administrativos",
                      "Uniformes + Materiais por período", COLORS[6])
    c5a.pack(side="left", fill="both", expand=True, padx=(0, 6))

    def kpi_adm():
        unif = [d.get("uniformes", 0) for d in dados]
        mat  = [d.get("materiais", 0) for d in dados]
        x = np.arange(len(meses))
        w = 0.38
        fig = _make_fig(plt, 6.0, 250)
        ax  = fig.add_subplot(111)
        ax.bar(x - w / 2, unif, w, label="Uniformes",       color=COLORS[6], alpha=0.9)
        ax.bar(x + w / 2, mat,  w, label="Mat. Escritório", color=COLORS[7], alpha=0.9)
        ax.set_xticks(x)
        ax.set_xticklabels(meses, fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: fmt_brl(v)))
        ax.tick_params(axis='y', labelsize=7)
        ax.legend(fontsize=8, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M)
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c5a, plt)

    kpi_adm()

    c5b = _card_frame(row5, "11. Índice HE por Funcionário",
                      "HE total ÷ headcount — sistêmico vs isolado", COLORS[4])
    c5b.pack(side="left", fill="both", expand=True)

    def kpi_he_func():
        idx   = [(d.get("he60_qtde", 0) + d.get("he100_qtde", 0))
                 / max(d.get("qtde_func", 1), 1) for d in dados]
        cores = [COLORS[4] if v > 0.5 else COLORS[2] for v in idx]
        fig = _make_fig(plt, 6.0, 250)
        ax  = fig.add_subplot(111)
        bars = ax.bar(meses, idx, color=cores, alpha=0.9)
        ax.axhline(0.5, color=WARNING, linestyle="--", linewidth=1.5, label="Ref. 0.5")
        for bar, val in zip(bars, idx):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 0.01, f"{val:.2f}",
                    ha="center", fontsize=7, color=FG_M)
        ax.set_xticklabels(meses, fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.2f}"))
        ax.tick_params(axis='y', labelsize=7)
        ax.legend(fontsize=8, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M)
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c5b, plt)

    kpi_he_func()

    # ── Rodapé ──
    rod = tk.Frame(inner, bg=BG)
    rod.pack(fill="x", padx=16, pady=(4, 14))
    tk.Label(rod,
             text=f"Fonte: {os.path.basename(caminho_arq)}  •  Aba: {aba_nome}  •  {len(meses)} período(s)",
             font=FONT_SMALL, bg=BG, fg=TEXT_SUB).pack(side="left")


# ══════════════════════════════════════════════════════════════════════════════
# COMPARATIVO POR LOJA
# ══════════════════════════════════════════════════════════════════════════════
def _janela_comparativo_lojas(root, caminho_arq, abas, cor):
    try:
        plt = _setup_mpl()
    except ImportError:
        messagebox.showerror("Dependência faltando", "matplotlib é necessário.")
        return

    dados_lojas = {}
    for aba in abas:
        try:
            d = _ler_ultimo_mes_aba(caminho_arq, aba)
            if d:
                dados_lojas[aba] = d
        except Exception:
            pass

    if not dados_lojas:
        messagebox.showwarning("Sem dados", "Não foi possível carregar dados das abas.")
        return

    lojas = list(dados_lojas.keys())
    n     = len(lojas)

    win = tk.Toplevel(root)
    win.title("Comparativo por Loja / Unidade")
    win.configure(bg=BG)
    win.geometry("1440x960")

    tk.Frame(win, bg="#065F46", height=5).pack(fill="x")
    top = tk.Frame(win, bg=BG)
    top.pack(fill="x", padx=28, pady=(14, 8))
    tk.Label(top, text="🏪", font=("Segoe UI", 22), bg=BG, fg="#065F46").pack(side="left", padx=(0, 10))
    col_t = tk.Frame(top, bg=BG)
    col_t.pack(side="left")
    tk.Label(col_t, text="Comparativo por Loja / Unidade",
             font=FONT_TITLE, bg=BG, fg=TEXT).pack(anchor="w")
    tk.Label(col_t, text=f"{n} unidade(s) — último mês disponível de cada aba",
             font=FONT_SMALL, bg=BG, fg=TEXT_SUB).pack(anchor="w")

    inner = _scroll_window(win)

    # ──────────────────────────────────────────────────
    # GRÁFICO DESTAQUE — Distribuição de Salários por Loja
    # Barras horizontais empilhadas: Bruto + Encargos + Benefícios
    # ──────────────────────────────────────────────────
    c_sal = _card_frame(inner,
                        "1. Distribuição de Salários por Loja / ADM",
                        "Composição: Salário Bruto + Encargos + Benefícios Líquidos — último mês de cada aba",
                        COLORS[0])
    c_sal.pack(fill="x", padx=16, pady=(10, 6))

    def graf_salarios_lojas():
        bruto    = [dados_lojas[l].get("bruto", 0)    for l in lojas]
        encargos = [dados_lojas[l].get("impostos", 0) for l in lojas]
        benef    = [dados_lojas[l].get("convenio", 0)
                    + max(dados_lojas[l].get("valor_vt", 0)
                          - dados_lojas[l].get("vt_desc_func", 0), 0)
                    for l in lojas]
        totais   = [b + e + bf for b, e, bf in zip(bruto, encargos, benef)]

        # Ordena por total decrescente
        ordem      = sorted(range(n), key=lambda i: totais[i], reverse=True)
        lojas_ord  = [lojas[i]    for i in ordem]
        bruto_ord  = [bruto[i]    for i in ordem]
        enc_ord    = [encargos[i] for i in ordem]
        benef_ord  = [benef[i]    for i in ordem]
        totais_ord = [totais[i]   for i in ordem]

        h_fig = max(320, n * 42)
        fig = _make_fig(plt, 13.5, h_fig)
        ax  = fig.add_subplot(111)

        y   = np.arange(len(lojas_ord))
        hb  = 0.55
        ax.barh(y, bruto_ord, hb, label="Salário Bruto",  color=COLORS[0], alpha=0.92)
        ax.barh(y, enc_ord,   hb, label="Encargos",        color=COLORS[4], alpha=0.92,
                left=bruto_ord)
        ax.barh(y, benef_ord, hb, label="Benefícios Liq.", color=COLORS[2], alpha=0.92,
                left=[a + b for a, b in zip(bruto_ord, enc_ord)])

        max_tot = max(totais_ord) if totais_ord else 1
        for i, tot_i in enumerate(totais_ord):
            ax.text(tot_i + max_tot * 0.005, i, fmt_brl(tot_i),
                    va="center", fontsize=8, color=FG_M, fontweight="bold")

        ax.set_yticks(y)
        ax.set_yticklabels(lojas_ord, fontsize=9)
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: fmt_brl(v)))
        ax.tick_params(axis='x', labelsize=7)
        ax.invert_yaxis()
        ax.legend(fontsize=8, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M,
                  loc="lower right")
        ax.grid(axis="x", alpha=0.4)
        ax.set_xlabel("R$", fontsize=8, color=FG_M)
        fig.tight_layout(pad=1.5)
        _embed(fig, c_sal, plt)

    graf_salarios_lojas()

    # ──────────────────────────────────────────────────
    # LINHA 2 — Custo por cabeça + Radar comparativo
    # ──────────────────────────────────────────────────
    row_b = tk.Frame(inner, bg=BG)
    row_b.pack(fill="x", padx=16, pady=6)

    c_cpc = _card_frame(row_b, "2. Custo por Cabeça por Loja",
                        "Total ÷ headcount — barras com linha de média geral", COLORS[2])
    c_cpc.pack(side="left", fill="both", expand=True, padx=(0, 6))

    def graf_cpc_lojas():
        cpc = []
        for l in lojas:
            d   = dados_lojas[l]
            tot = sum(v for k, v in d.items() if isinstance(v, float) and "qtde" not in k)
            q   = max(d.get("qtde_func", 1), 1)
            cpc.append(tot / q)
        media = sum(cpc) / len(cpc)
        cores = [COLORS[i % len(COLORS)] for i in range(len(lojas))]
        fig = _make_fig(plt, 6.5, 280)
        ax  = fig.add_subplot(111)
        bars = ax.bar(lojas, cpc, color=cores, alpha=0.9)
        ax.axhline(media, color=WARNING, linestyle="--", linewidth=1.5,
                   label=f"Média {fmt_brl(media)}")
        for bar, val in zip(bars, cpc):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + max(cpc) * 0.01,
                    fmt_brl(val), ha="center", fontsize=7, color=FG_M)
        ax.set_xticklabels(lojas, fontsize=8, rotation=20, ha="right")
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: fmt_brl(v)))
        ax.tick_params(axis='y', labelsize=7)
        ax.legend(fontsize=8, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M)
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c_cpc, plt)

    graf_cpc_lojas()

    c_rad = _card_frame(row_b, "3. Radar Comparativo de Saúde",
                        "Multidimensional por unidade — até 5 lojas visíveis", COLORS[5])
    c_rad.pack(side="left", fill="both", expand=True)

    def graf_radar_lojas():
        dims   = ["Turnover\nCtrl", "Efic. HE", "Recup.", "Custo\nBenef", "Adm Ctrl"]
        N      = len(dims)
        angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist() + [0]
        fig = _make_fig(plt, 5.8, 280)
        ax  = fig.add_subplot(111, polar=True)
        ax.set_facecolor("#1C1C28")
        for idx_l, loja in enumerate(lojas[:5]):
            d   = dados_lojas[loja]
            b   = max(d.get("bruto", 1), 1)
            q   = max(d.get("qtde_func", 1), 1)
            tot = sum(v for k, v in d.items() if isinstance(v, float) and "qtde" not in k) or 1
            vals = [
                max(0, 100 - (d.get("rescisao", 0) / b * 100) * 5),
                max(0, 100 - (d.get("he_total_qtde", 0) / q) * 80),
                min(100, (d.get("vt_desc_func", 0) + d.get("refeicoes_desc", 0)) / b * 1000),
                min(100, max(0, 100 - (d.get("convenio", 0) + d.get("valor_vt", 0)) / tot * 200)),
                min(100, max(0, 100 - (d.get("uniformes", 0) + d.get("materiais", 0)) / tot * 300)),
            ]
            vals_p = vals + vals[:1]
            clr    = COLORS[idx_l % len(COLORS)]
            ax.fill(angles, vals_p, color=clr, alpha=0.12)
            ax.plot(angles, vals_p, color=clr, linewidth=1.8,
                    marker="o", markersize=4, label=loja)
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(dims, fontsize=8, color=FG_M)
        ax.set_ylim(0, 100)
        ax.set_yticks([25, 50, 75, 100])
        ax.set_yticklabels(["25", "50", "75", "100"], fontsize=6, color=SUB_M)
        ax.spines["polar"].set_color("#2A2A3E")
        ax.grid(color="#2A2A3E", alpha=0.6)
        ax.legend(fontsize=7.5, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M,
                  loc="upper right", bbox_to_anchor=(1.45, 1.15))
        fig.patch.set_facecolor(BG_M)
        fig.tight_layout(pad=1.5)
        _embed(fig, c_rad, plt)

    graf_radar_lojas()

    # ──────────────────────────────────────────────────
    # LINHA 3 — HE por loja + Turnover por loja
    # ──────────────────────────────────────────────────
    row_c = tk.Frame(inner, bg=BG)
    row_c.pack(fill="x", padx=16, pady=6)

    c_he = _card_frame(row_c, "4. HE por Loja",
                       "Índice HE ÷ headcount — quem está sobrecarregado", COLORS[4])
    c_he.pack(side="left", fill="both", expand=True, padx=(0, 6))

    def graf_he_lojas():
        idx_he = [(dados_lojas[l].get("he60_qtde", 0) + dados_lojas[l].get("he100_qtde", 0))
                  / max(dados_lojas[l].get("qtde_func", 1), 1) for l in lojas]
        cores  = [COLORS[4] if v > 0.5 else COLORS[2] for v in idx_he]
        fig = _make_fig(plt, 6.5, 260)
        ax  = fig.add_subplot(111)
        bars = ax.bar(lojas, idx_he, color=cores, alpha=0.9)
        ax.axhline(0.5, color=WARNING, linestyle="--", linewidth=1.5, label="Ref. 0.5 HE/func")
        for bar, val in zip(bars, idx_he):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 0.01, f"{val:.2f}",
                    ha="center", fontsize=7, color=FG_M)
        ax.set_xticklabels(lojas, fontsize=8, rotation=20, ha="right")
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.2f}"))
        ax.tick_params(axis='y', labelsize=7)
        ax.legend(fontsize=8, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M)
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c_he, plt)

    graf_he_lojas()

    c_turn = _card_frame(row_c, "5. Turnover por Loja",
                         "Rescisões / Bruto — identifica unidades em alerta", COLORS[4])
    c_turn.pack(side="left", fill="both", expand=True)

    def graf_turnover_lojas():
        ratio = [(dados_lojas[l].get("rescisao", 0)
                  / max(dados_lojas[l].get("bruto", 1), 1)) * 100 for l in lojas]
        cores = [COLORS[4] if r > 10 else COLORS[2] for r in ratio]
        fig = _make_fig(plt, 6.5, 260)
        ax  = fig.add_subplot(111)
        bars = ax.bar(lojas, ratio, color=cores, alpha=0.9)
        ax.axhline(10, color=WARNING, linestyle="--", linewidth=1.5, label="Limite 10%")
        for bar, val in zip(bars, ratio):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 0.2, f"{val:.1f}%",
                    ha="center", fontsize=7, color=FG_M)
        ax.set_xticklabels(lojas, fontsize=8, rotation=20, ha="right")
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.1f}%"))
        ax.tick_params(axis='y', labelsize=7)
        ax.legend(fontsize=8, facecolor=BG_M, edgecolor=BG_M, labelcolor=FG_M)
        ax.grid(axis="y")
        fig.tight_layout(pad=1.5)
        _embed(fig, c_turn, plt)

    graf_turnover_lojas()

    # ── Rodapé ──
    rod = tk.Frame(inner, bg=BG)
    rod.pack(fill="x", padx=16, pady=(4, 14))
    tk.Label(rod,
             text=f"Fonte: {os.path.basename(caminho_arq)}  •  {n} unidade(s) comparadas  •  último mês de cada aba",
             font=FONT_SMALL, bg=BG, fg=TEXT_SUB).pack(side="left")
    


def tela_excel_despesa(root):
    def corpo(root, content, cor):
        caminho = tk.StringVar(value="")

        # Card de upload
        card = tk.Frame(content, bg=SURFACE, padx=24, pady=20)
        card.pack(fill="x", pady=(0, 14))

        tk.Label(card, text="Arquivo Excel", font=FONT_HEAD, bg=SURFACE, fg=TEXT).pack(anchor="w")
        tk.Label(card, text="Formatos suportados: .xlsx", font=FONT_SMALL, bg=SURFACE, fg=TEXT_SUB).pack(anchor="w", pady=(2, 12))

        row = tk.Frame(card, bg=SURFACE)
        row.pack(fill="x")

        entry = tk.Entry(row, textvariable=caminho, bg=SURFACE2, fg=TEXT,
                         insertbackground=TEXT, relief="flat",
                         font=FONT_MONO, highlightthickness=1,
                         highlightbackground=BORDER, highlightcolor=cor)
        entry.pack(side="left", fill="x", expand=True, ipady=8, padx=(0, 10))
        entry.insert(0, "Nenhum arquivo selecionado...")

        def selecionar():
            p = filedialog.askopenfilename(
                title="Selecione o arquivo Excel",
                filetypes=[("Excel", "*.xlsx *.xls")]
            )
            if p:
                caminho.set(p)

        pill_button(row, "Navegar", selecionar, color=cor, hover=ACCENT2).pack(side="right")

        # Opções
        card2 = tk.Frame(content, bg=SURFACE, padx=24, pady=16)
        card2.pack(fill="x", pady=(0, 14))
        tk.Label(card2, text="Opções de Processamento", font=FONT_HEAD, bg=SURFACE, fg=TEXT).pack(anchor="w", pady=(0, 10))

        opts = [("Remover duplicatas", True), ("Formatar cabeçalhos", True), ("Exportar relatório", False)]
        vars_ = []
        for label, default in opts:
            v = tk.BooleanVar(value=default)
            vars_.append(v)
            row_opt = tk.Frame(card2, bg=SURFACE)
            row_opt.pack(fill="x", pady=2)
            tk.Checkbutton(row_opt, variable=v, bg=SURFACE, fg=TEXT,
                           selectcolor=cor, activebackground=SURFACE,
                           font=FONT_BODY).pack(side="left")
            tk.Label(row_opt, text=label, bg=SURFACE, fg=TEXT, font=FONT_BODY).pack(side="left")

        def executar():
            arq = caminho.get()
            if not arq or "Nenhum" in arq:
                messagebox.showwarning("Atenção", "Selecione um arquivo antes de executar.")
                return
            
            try:
                
                from services.despesas.main import iniciar_processamento
                
                iniciar_processamento(arq)
                
                messagebox.showinfo("Sucesso", "Processamento concluído!")
            except Exception as e:
                messagebox.showerror("Erro de Módulo", f"Não foi possível iniciar o módulo: {e}")

        pill_button(content, "▶  Executar Automação", executar, color=cor, hover=ACCENT2).pack(fill="x", pady=4, ipady=4)

    _base_tela(root, "Processar Excel", "📊", ACCENT, corpo)
